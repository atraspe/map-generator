# Import load_workbook from openpyxl module to handle how the maps specs is parsed (usually an Excel spreadsheet)
from openpyxl import load_workbook

# Import deque from collections module in order to store rules for the source model (queue data structure)
from collections import deque

# Import Path to work with files
from pathlib import Path

# For datetime stamp for header comments
from datetime import datetime

# Load entire spreadsheet/workbook into this object
wb = load_workbook('Example_Spec_v1.0.xlsx')

# Print all workbook sheets
print(f'Sheet names:{wb.sheetnames}\n')

# Obtain information from Title sheet
title_sheet = wb['Title']

# Build the map name based on information from Title Sheet
# Map name convention: 5-digit client ID + direction + document type + document version + s(source)/t(target) + .mdl
for i in range(2, 11):
    for j in range(1, 3):
        if j == 2:
            key = title_sheet.cell(row=i, column=j-1).value.lower()
            value = title_sheet.cell(row=i, column=j).value
            if key[:18] == 'client solution id':
                client_id = value
            if key[:15] == 'trading partner':
                tp_name = value
            if key[:15] == 'edi transaction':
                doc_type = value
            if key[:11] == 'edi version':
                doc_version = value
            if key[:9] == 'direction':
                direction = value
                dir_io = 'o' if value.lower() == 'outbound' else 'i'
                s_or_t = 's' if value.lower() == 'outbound' else 't'
        

map_name = client_id + dir_io + str(doc_type) + str(doc_version) + s_or_t + '.mdl'
new_map = Path(map_name)
print(f'Map file exists?: {new_map.exists()}')

# Obtain information from UDF (User-Defined Format) sheet
udf_sheet = wb['UDF']
udf_max_row = udf_sheet.max_row
udf_max_col = udf_sheet.max_column

# print(f'Max row: {udf_max_row}')
# print(f'Max col: {udf_max_col}')

def get_cell_value(sheet, row, col):
    return sheet.cell(row=row, column=col).value

# To be able to manage potentially large specs, each row will be stored in a generator object
all_UDF_rows = (
    (get_cell_value(udf_sheet, i, j) for j in range(1, udf_max_col + 1)) for i in range(2, udf_max_row + 1)
    )


# How information from each row will be parsed:
#   - each item will be stored in a dictionary
#   - key will be the data model item name; value will be a list (of dictionaries) of the item's properties
#   - each parent item will be noted - this is necessary to determine when to correctly close each parent (Group or Record)
#   - e.g. [ {'Initialization' : [ {'Occ Min' : 1}, {'Occ Max' : 9999}, {'Rules' : ['1st rule', '2nd rule', 'nth rule'] } ] } ]

#   - 'Rules' key will be using a queue value - contains every field's/element's mapping logic/rules
#       - each item will be appended using .append() while retrieved using .popleft()
#       - all these ARRAYs will be written in the PRESENT rules section of their parent item
#       - e.g. ARRAY->Field1 = STRTRIM(DEFAULT_NULL(&Field1), VAR->TRIM_TYPE, VAR->SPACE)
parent_items = []


# The below list will contain every data model item's closing loop
# e.g. }*0 .. 1 ;; |-- end F57_FILLER_1 --|
#      }*0 .. * ;; |-- end F16_GROUP --|
#      }*1 .. 1 ;; |-- end Document_Loop --|
closing_loops = []


# These are fixed-length fields from OTFixed.acc
# fixed_length_fields = ['RptLineFld', 'AlphaFld', 'CsvFld', 'AlphaNumericFld', 'AlphaNumericReqFld', 'DateFld', 'DateFldNA', 'TimeFld', 
#                     'TimeFldNA', 'NumericFld', 'NumericFldNA', 'IDFld', 'AnyCharDelim', 'AnyCharDelimCRLF', 'AnyCharODelim', 
#                     'AnyCharDelimNCR', 'AnyCharDelimCR', 'QuestionMark']
flf_list = ['AN', 'DT', 'NF', 'TF']
flf_dict = {'AN' : 'AlphaNumericFld', 'DT' : 'DateFld', 'NF' : 'NumericFld', 'TF' : 'TimeFld'}

# These are fixed-length records from OTFixed.acc
# fixed_length_records = ['Rec_Code', 'FixedLgthRecord', 'FixedLgthDefaultRecord', 'LineFeedDelimRecord', 'LineFeedDelimDefaultRecord', 
#                     'LineFeedDelimContainer', 'LineFeedDelimRecordCR', 'VariableLgthRecord', 'Group']
flr_list = ['LS', 'LE', 'RC']
flr_dict = {'LS' : '', 'LE' : '', 'RC' : 'LineFeedDelimRecord'}

# These are data model item types from the specs
data_model_item_types = ['AN', 'DT', 'NF', 'LE', 'LS', 'RC', 'RT', 'TF']

# Global variables which will be used by DataModelItem() class
prev_item_type = ''
opening_bracket, closing_bracket = '{', '}'
previous_occ_max, item_counter = 0, 0

# This list will hold all warning and error messages during specs validation process
warnings = []


class DataModelItem():
    # This is the main class which does the following:
    #   - unpacks the tuple of attributes taken from each row

    def __init__(self, field_tuple):
        # Unpack the tuple of attributes
        self.loop_rec_name, self.field_name, self.desc, self.type, self.dec, self.format, \
        self.start, self.end, self.length, self.occurence, self.mapped, self.correlation, self.comments, self.spec_logic = field_tuple
        
        self.occ_min, self.occ_max = (0, 0)
        
        if self.type in ['LS', 'RC']:
            # For Loop start (LS) and Record (RC) items,
            #   the Occ value comes in as x .. y so need to separate the min(x) from the max(y)
            occ = self.occurence
            self.occ_min = occ.split('..')[0].strip()
            self.occ_max = occ.split('..')[-1].strip()
                
        elif self.type in ['RT', 'AN']:
            # For Record tag (RT) and fixed-length fields (AlphaNumeric, Numeric, etc),
            #   the Occ value comes in as M (mandatory) or O (optional)
            if self.occurence == 'M':
                self.occ_min, self.occ_max = (1, 1)
            else:
                self.occ_min, self.occ_max = (0, 1)
                
        elif self.type == 'LE':
            # Loop end (LE) do not have Occ value
            #   so get their parent item's Occurence min and max values from the parent_items[]
            self.occ_min = list(parent_items[-1].values())[0][0]['Occ Min']
            self.occ_max = list(parent_items[-1].values())[0][1]['Occ Max']


    def parse_item(self):
        # This function will prepare some values prior to calling a private method _generate_item()
        global item_counter

        # Determine data model item name to be used
        dmi_name = self.loop_rec_name if self.type in flr_list else self.field_name

        format = f' "{self.format}"' if self.format else ''
        length = f' @{self.length} .. {self.length}'

        generated_item = self._generate_item(dmi_name, format, length)
        map_file.write(generated_item)


        if self.type != 'RC':
            # If type is 'RC", the expectation is that next item will be a 'RT' (Record Tag)
            # so next iteration would just write the value from Correlation column of the specs
            map_file.write('\n')
        
        if item_counter == 0:
            # For the first item (Initialization), map documentation header will be generated
            header_comments = self._generate_header_comments()
            list(parent_items[-1].values())[-1][2]["Rules"].append(header_comments)
            mapping_rules = list(parent_items[-1].values())[-1][2]["Rules"]
            self._write_mapping_rules(mapping_rules)
            item_counter += 1


    def _generate_header_comments(self):
        global client_id, map_name, direction, doc_type, doc_version, tp_name

        header_comments = f''';;
;; o------------------------------------------------------------------------------------------o
;; |                   Map Documentation
;; o------------------------------------------------------------------------------------------o
;; |         Client Name: {client_id}
;; |             Program: {map_name}
;; |           Direction: {direction}
;; |            Standard: UDF
;; |            Document: {doc_type}
;; |             Version: {doc_version}
;; |     Trading Partner: {tp_name}
;; |        Developed By: floresl
;; |      Date Developed: {datetime.now().strftime("%c")}
;; |    Last Modified By:
;; |  Date Last Modified:
;; o------------------------------------------------------------------------------------------o

[ ]
VAR->NULL = ""
VAR->TRIM_TYPE = "B"
VAR->SPACE = " "
VAR->DATA = "Data"
VAR->STOP = "Stop"
VAR->YES = "Yes"
VAR->NO = "No"

[ ]
VAR->Workbench = VAR->YES
VAR->Session = VAR->NULL
VAR->Session = VAR->OTSessionNo

[VAR->Session != VAR->NULL]
VAR->Workbench = VAR->NO

[VAR->Workbench == VAR->YES]
PERFORM("OTSessionInit")

[ ]
;;; ECSC Standard Map PERFORM
PERFORM ("ECSCOutbSourceInit")
PERFORM ("OTAdminInit")
'''
        return header_comments



    def _write_mapping_rules(self, mapping_rules):
        # A simple private method to write all the mapping rules in a parent's PRESENT rules section
        map_file.write('[ ]\n')
        while mapping_rules:
            map_file.write(f'{mapping_rules.popleft()}\n')


    def _generate_item(self, dmi_name, format, length_min_max):
        # The beginning of the item will be generated, but will wait for the next iteration (next item)
        #   to determine if it'll close this item (build the trailer) or build the next one (if it's a child of this current item)
        global parent_items, closing_loops

        # A lambda to construct data model item's leading part
        #   - 1st parameter is a boolean value to determine if it's a parent
        #   - 2nd parameter is the data model item name
        item_leading = lambda parent, dmi_name : f'{dmi_name} {opening_bracket} {flr_dict[self.type] if parent else flf_dict[self.type]}{"" if parent else length_min_max}{"" if parent else format} {"" if parent else "none"}'

        # A lambda to construct data model item's trailing part
        item_trailing = lambda dmi_name : f'{closing_bracket}*{self.occ_min} .. {self.occ_max} ;; |-- end {dmi_name} --|'

        # A lambda to create the mapping rule to store field's value into ARRAY
        rule_array = lambda dmi_name : f'ARRAY->{dmi_name} = STRTRIM(DEFAULT_NULL(&{dmi_name}), VAR->TRIM_TYPE, VAR->SPACE)'

        if self.type in data_model_item_types:
            if self.type == 'RT':
                # If it's a Record Tag, just return the value
                return self.correlation
            
            # Check first if need to close out prior data model item, that is if:
            #   a) current item is a sibling of previous item, or
            #   b) current item is another record, group or if it's the end of the data model

            if (self.type in flf_list and prev_item_type in flf_list) or (self.type in flr_list and prev_item_type in flf_list):
                # Current item is a fixed length field and previous is also a field, OR
                # Current item is a Group (LS) or start of another record (RE), then
                # 1) Close out the previous item's trailing
                map_file.write(f'{closing_loops.pop()}\n')
                
            if self.type in flr_list and prev_item_type in flf_list:
                # Current item is a Group (LS) or start of another record (RE), then
                # 1) Write all the mapping rules in the PRESENT section of parent item
                mapping_rules = list(parent_items[-1].values())[-1][2]["Rules"]
                self._write_mapping_rules(mapping_rules)

                # 2) Close out the previous item's trailing (essentially closing out the last parent item)
                map_file.write(f'{closing_loops.pop()}\n')

                # 3) Pop the last item in parent_items list (take off from the list)
                parent_items.pop()

            
            # Save the current item's trailing
            closing_loops.append(item_trailing(dmi_name))

            if self.type == 'LS' or self.type == 'RC':
                # 'LS (Loop Start)' and 'RC (Record)' will be treated as parent items
                # Save data model item in parent_items
                parent_items.append({self.loop_rec_name : [
                    {'Occ Min' : self.occ_min},
                    {'Occ Max' : self.occ_max},
                    {'Rules' : deque([])}
                    ]})
                
                # Save the current item's trailing
                closing_loops.append(item_trailing(dmi_name))
                return item_leading(True, dmi_name)
            
            if self.type == 'LE':
                loop_end = dmi_name[4:]
                last_parent = list(parent_items[-1].keys())[-1]

                if loop_end.lower() == last_parent.lower():
                    # Pop the last item in parent_items list
                    parent_items.pop()
                    # Close the item, return its trailer
                    return item_trailing(loop_end)
                            
            if self.type in flf_list:
                # Append this data model item's map logic into the mapping logic queue
                list(parent_items[-1].values())[-1][2]["Rules"].append(rule_array(dmi_name))
                return item_leading(False, dmi_name)
        

        else:
            return 'Invalid Data Model Item Type!'

# Create the map
with open(new_map, 'w', newline=None) as map_file:
    for attributes in all_UDF_rows:
        # Convert each row's attributes into a tuple and pass to the DataModelItem class in an expected order
        new_DMI = DataModelItem(tuple(attributes))
        new_DMI.parse_item()
                
        # Save current data model item type for comparison with the next item
        prev_item_type = new_DMI.type

        # Delete current object prior to next loop iteration
        del new_DMI
